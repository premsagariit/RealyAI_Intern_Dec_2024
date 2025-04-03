import os
import time
import hashlib
import requests
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

# API Details
url_pre = "https://ap-east-1.tensorart.cloud"
api_key = "38dbb245-327b-485c-bfdf-63ae966edb73"  # Replace with your actual API key
url_job = "/v1/jobs"
url_resource = "/v1/resource"
url_workflow = "/v1/workflows"

# Headers for API Key Authentication
HEADERS = {
    'Content-Type': 'application/json',
    'Accept': 'application/json',
    'Authorization': f'Bearer {api_key}'
}

def ensure_output_folder():
    """Ensure that the 'outputs' folder exists."""
    if not os.path.exists("outputs"):
        os.makedirs("outputs")

def save_image(image_url):
    """Download and save the generated image. Return the saved image path."""
    ensure_output_folder()
    response = requests.get(image_url)
    if response.status_code == 200:
        # Create a unique filename based on a hash of the URL
        image_path = os.path.join("outputs", f"{hashlib.md5(image_url.encode()).hexdigest()}.png")
        with open(image_path, "wb") as img_file:
            img_file.write(response.content)
        print(f"Image saved to {image_path}")
        return image_path
    else:
        print("Error downloading image.")
        return None

def get_job_result(job_id):
    """Poll the API for job completion status and return the output image path when complete."""
    while True:
        time.sleep(1)
        response = requests.get(f"{url_pre}{url_job}/{job_id}", headers=HEADERS)
        job_response = response.json()
        if 'job' in job_response:
            job_dict = job_response['job']
            job_status = job_dict.get('status')
            print("Job status:", job_dict)
            if job_status in ['SUCCESS', 'FAILED']:
                if job_status == 'SUCCESS':
                    image_url = job_dict['successInfo']['images'][0]['url']
                    print(f"Image URL: {image_url}")
                    print("Image generation successful.")
                    return save_image(image_url)
                else:
                    print("Image generation failed.")
                    return None

def upload_img(img_path):
    """Upload an image and return the resource ID."""
    print(f"Uploading image: {img_path}")
    data = {"expireSec": 3600}
    response = requests.post(f"{url_pre}{url_resource}/image", json=data, headers=HEADERS)
    print("Upload response:")
    print(response.text)
    response_data = response.json()
    resource_id = response_data['resourceId']
    put_url = response_data['putUrl']
    headers = response_data['headers']
    
    with open(img_path, 'rb') as f:
        res = f.read()
        upload_response = requests.put(put_url, data=res, headers=headers)
        print("PUT response:")
        print(upload_response.text)
    
    return resource_id

def generate_image(resource_id, positive_prompt, style, negative_prompt):
    """
    Submit a job using the predefined workflow template with template ID 688362427502551075.
    The style (and prompt) are provided as parameters.
    """
    data = {
        "request_id": hashlib.md5(str(int(time.time())).encode()).hexdigest(),
        "templateId": "688362427502551075",
        "params": {
            "async": False,
            "priority": "NORMAL",
            "extraParams": {}
        },
        "fields": {
            "fieldAttrs": [
                # Node 12: LoadImage
                {"nodeId": "12", "fieldName": "image", "fieldValue": resource_id},
                # Node 25: SDXL Prompt Styler
                {"nodeId": "25", "fieldName": "style", "fieldValue": style},
                {"nodeId": "25", "fieldName": "text_positive", "fieldValue": positive_prompt},
                {"nodeId": "25", "fieldName": "text_negative", "fieldValue": negative_prompt},
                # Node 18: IPAdapter FaceID
                {"nodeId": "18", "fieldName": "weight", "fieldValue": "1.2"},
                {"nodeId": "18", "fieldName": "weight_faceidv2", "fieldValue": "1.5"},
                {"nodeId": "18", "fieldName": "combine_embeds", "fieldValue": "average"},
                {"nodeId": "18", "fieldName": "start_at", "fieldValue": "0.150"},
                {"nodeId": "18", "fieldName": "end_at", "fieldValue": "1.000"},
                {"nodeId": "18", "fieldName": "embeds_scaling", "fieldValue": "K+V w/ C penalty"},
                # Node 3: KSampler
                {"nodeId": "3", "fieldName": "seed", "fieldValue": "161098661698898"},
                {"nodeId": "3", "fieldName": "cfg", "fieldValue": "2"},
                {"nodeId": "3", "fieldName": "steps", "fieldValue": "20"},
                {"nodeId": "3", "fieldName": "sampler_name", "fieldValue": "dpmpp_2m"},
                {"nodeId": "3", "fieldName": "scheduler", "fieldValue": "sgm_uniform"},
                {"nodeId": "3", "fieldName": "control_after_generate", "fieldValue": "fixed"},
                
                # Node 5: Empty Latent Image
                {"nodeId": "5", "fieldName": "width", "fieldValue": "1024"},
                {"nodeId": "5", "fieldName": "height", "fieldValue": "1024"},
                # Node 4: Load Checkpoint
                {"nodeId": "4", "fieldName": "ckpt_name", "fieldValue": "676746065682318967"},
                # Node 38: LoraLoaderModelOnly
                {"nodeId": "38", "fieldName": "lora_name", "fieldValue": "681174344216573550"},
                {"nodeId": "38", "fieldName": "strength_model", "fieldValue": "0.7"},
                {"nodeId": "38", "fieldName": "tams_lora_name", "fieldValue": "IP_adapter_face_id - SDXL"}
            ]
        },
    }
    
    # Submit the job
    response = requests.post(f"{url_pre}{url_job}/workflow/template", json=data, headers=HEADERS)
    print("Template job submission response:")
    print(response.text)
    response_data = response.json()
    
    if 'job' in response_data:
        job_dict = response_data['job']
        job_id = job_dict.get('id')
        print(f"Workflow Job ID: {job_id}, Status: {job_dict.get('status')}")
        # Poll until the job is complete and then return the output image path
        return get_job_result(job_id)
    else:
        print("Failed to create workflow job:", response_data)
        return None

# --- Helper functions for Excel cell dimension adjustments ---

def pixels_to_points(pixels):
    """Convert pixel height to Excel row height in points (rough conversion)."""
    return pixels * 0.75

def pixels_to_excel_width(pixels):
    """Convert pixel width to Excel column width (rough conversion)."""
    return pixels * 0.14

def main():
    # Folder with input images
    input_folder = "Inputs"
    # Get list of image files (adjust extensions as needed)
    input_images = [os.path.join(input_folder, f) for f in os.listdir(input_folder)
                    if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    
    if not input_images:
        print("No input images found in the 'Inputs' folder.")
        return

    # Get prompt inputs (or these could be hardcoded)
    positive_prompt = input("Enter the prompt for the images: ")
    negative_prompt = ("lowres, bad hands, text, error, missing fingers, extra digit, fewer digits, "
                         "cropped, worst quality, low quality, normal quality, jpeg artifacts, signature, "
                         "watermark, username, blurry, patreon logo, artist name, sexual content, adult, nudity")
    
    # Define your list of styles (modify as needed)

    styles_list = ["sai-comic book", "artstyle-watercolor"]
    # styles_list = ["sai-comic book", "sai-line art", "artstyle-pop art", "artstyle-watercolor", "misc-kawaii"]
    # Create an Excel workbook and set up the header row
    wb = Workbook()
    ws = wb.active
    ws.title = "TensorART Results"
    ws['A1'] = "Input Image"
    ws['B1'] = "Style"
    ws['C1'] = "Output Image"
    
    current_excel_row = 2  # start writing from row 2
    max_width_A = 0  # track maximum input image width in pixels (scaled)
    max_width_C = 0  # track maximum output image width in pixels (scaled)

    # Define maximum allowed pixel dimension for 5cm (assuming 96 DPI, 5cm ≈ 189 pixels)
    max_pixels = 189

    # Loop over each input image
    for input_img in input_images:
        print(f"Processing input image: {input_img}")
        # Upload the input image once to get the resource ID
        resource_id = upload_img(input_img)
        
        # Get input image dimensions using Pillow
        pil_input = PILImage.open(input_img)
        input_width, input_height = pil_input.size
        # Calculate scale factor so neither dimension exceeds max_pixels while preserving aspect ratio
        scale_input = min(1, max_pixels / input_width, max_pixels / input_height)
        scaled_input_width = int(input_width * scale_input)
        scaled_input_height = int(input_height * scale_input)
        
        # Loop over each style for the current input image
        for style in styles_list:
            print(f"Processing style: {style} for image: {input_img}")
            # Generate the image with the given style
            output_image_path = generate_image(resource_id, positive_prompt, style, negative_prompt)
            # Wait for 5 seconds after each API call
            time.sleep(5)
            
            # --- Embed the input image (scaled) into the Excel sheet ---
            img_excel_input = ExcelImage(input_img)
            # Set the scaled dimensions for the Excel image
            img_excel_input.width = scaled_input_width
            img_excel_input.height = scaled_input_height
            cell_A = f"A{current_excel_row}"
            ws.add_image(img_excel_input, cell_A)
            
            # Insert the style text in column B
            ws[f"B{current_excel_row}"] = style
            
            # --- Process the output image ---
            if output_image_path:
                pil_output = PILImage.open(output_image_path)
                out_width, out_height = pil_output.size
                scale_output = min(1, max_pixels / out_width, max_pixels / out_height)
                scaled_out_width = int(out_width * scale_output)
                scaled_out_height = int(out_height * scale_output)
                
                img_excel_output = ExcelImage(output_image_path)
                img_excel_output.width = scaled_out_width
                img_excel_output.height = scaled_out_height
                cell_C = f"C{current_excel_row}"
                ws.add_image(img_excel_output, cell_C)
            else:
                scaled_out_height = 0
                scaled_out_width = 0
            
            # Update maximum width trackers (for adjusting column widths later)
            max_width_A = max(max_width_A, scaled_input_width)
            max_width_C = max(max_width_C, scaled_out_width)
            
            # Set the row height to accommodate the taller of the two images (converted to points)
            row_height = pixels_to_points(max(scaled_input_height, scaled_out_height))
            ws.row_dimensions[current_excel_row].height = row_height
            
            current_excel_row += 1
    
    # After processing all rows, adjust the column widths.
    ws.column_dimensions['A'].width = pixels_to_excel_width(max_width_A)
    ws.column_dimensions['C'].width = pixels_to_excel_width(max_width_C)
    # For column B (style text), adjust width based on the longest style string
    max_style_length = max(len(style) for style in styles_list)
    ws.column_dimensions['B'].width = max_style_length * 1.2  # rough factor
    
    # Save the Excel workbook.
    # output_excel = "res.xlsx"
    output_excel = "results1.xlsx"
    wb.save(output_excel)
    print(f"Excel file saved as {output_excel}")

if __name__ == '__main__':
    main()
